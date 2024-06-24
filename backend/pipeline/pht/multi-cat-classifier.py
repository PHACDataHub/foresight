from collections import defaultdict
import configparser
import sys

from openpyxl import load_workbook, Workbook

from langchain.output_parsers.list import NumberedListOutputParser
from langchain.prompts.prompt import PromptTemplate
# from langchain.text_splitter import RecursiveCharacterTextSplitter
# from langchain.chains.summarize import load_summarize_chain

from langchain_community.llms import Ollama


def load_config(file_name):
    config = configparser.ConfigParser()
    config.read_file(open(file_name))
    return config


def print_category_dict(category_dict):
    for k1, v1 in category_dict.items():
        print(f"{k1}")
        if isinstance(v1, dict):
            for k2, v2 in v1.items():
                print(f"\t{k2}")
                for e in v2:
                    print(f"\t\t{e}")
        elif isinstance(v1, list):
            for e in v1:
                print(f"\t{e}")
        else:
            raise Exception(f"Unknown category type -- {v1}")
    return category_dict


def load_categories(config):
    category_dict = eval(config['public_health_threats']['THREAT_CATEGORIES'])
    # print_category_dict(category_dict)
    return category_dict


def match_category(category_dict, category):
    for k1, v1 in category_dict.items():
        if k1.lower() == category.lower():
            return k1
        if isinstance(v1, dict):
            for k2, v2 in v1.items():
                if k2.lower() == category.lower():
                    return k2
                for e in v2:
                    if e.lower() == category.lower():
                        return e
        else:   # if isinstance(v1, list):
            for e in v1:
                if e.lower() == category.lower():
                    return e
    raise Exception(f"Unknown category -- {category}")


def load_data_file(file_name, sheet_name, category_dict):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    headers, inputs = [], []
    for i, row in enumerate(ws): 
        if i == 0:
            headers = [cell.value for cell in row]
            continue
        row_dict = {headers[j]: cell.value for j, cell in enumerate(row)}

        categories = [e.strip().strip("'") for e in row_dict['ground_truth'].strip('[]').split(',') if e ]
        print(categories)
        truths = [match_category(category_dict, c) for c in categories]
        print(truths)
        row_dict['ground_truth'] = truths

        inputs.append(row_dict)
    return inputs


def write_to_excel_file(file_name, sheet_name, outputs):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(outputs[0].keys())
    ws.append(headers)
    for output in outputs:
        row = [f"{output[h]}" for h in headers]
        ws.append(row)
    wb.save(file_name)


def create_summarizer_prompt(format_instructions):
    MAP_PROMPT = """
    Return a concise summary of the following article focusing on factors impacting public health risks, highlighting possible dangers and threats:
    {article}

    {format_instructions}"""
    return PromptTemplate(
        input_variables=["article"],
        partial_variables={"format_instructions": format_instructions}, 
        template=MAP_PROMPT)


def summarize_text(model_name, llm_chain, output):
    result = llm_chain.invoke({"article": output['text']})
    if result and isinstance(result, list) and result[0]:
        output['sum'] = result[0]

    print(f"[{model_name} SUMMARIZER] --- {output['url']} \n\t--- {output['sum']}\n")
    return output


def create_classifier_prompt(format_instructions):
    CLASSIFY_PROMPT_TEMPLATE = """
    Which hazard listed below:
    {hazards}

    that best matches the article delimited by triple backquotes (```):
    ```{article}```

    If there is no such topic, then return No match as result.
    {format_instructions}"""
    return PromptTemplate(
        input_variables=["article", "hazards"],
        partial_variables={"format_instructions": format_instructions}, 
        template=CLASSIFY_PROMPT_TEMPLATE)


def classify_text(llm_chain, text, topics):
    identified_topics = []
    results = llm_chain.invoke({"article": text, "hazards": topics})
    for result in results:
        for topic in topics:
            if result in topic or topic in result:
                identified_topics.append(topic)
                break
    return identified_topics


def classify_threats(model_name, llm_chain, category_dict, output):
    print(f"[{model_name} CLASSIFIER] --- {output['url']} --- {output['ground_truth']}\n")
    
    for prp, hdr in zip(['text', 'sum'], ['full_text_top_down_threats', 'summary_top_down_threats']):
        text = output[prp]
        cls_threats = output[hdr]
        l1_threats = classify_text(llm_chain, text, list(category_dict.keys()))
        for l1_threat in l1_threats:
            cls_threats.append(l1_threat)
            if isinstance(category_dict[l1_threat], dict):
                l2_threats = classify_text(llm_chain, text, category_dict[l1_threat].keys())
                for l2_threat in l2_threats:
                    cls_threats.append(l2_threat)
                    cls_threats.extend(classify_text(llm_chain, text, category_dict[l1_threat][l2_threat]))
            else:
                cls_threats.extend(classify_text(llm_chain, text, category_dict[l1_threat]))

        print(f"\t[{model_name} {prp} {hdr}] --- {output['url']} \n\t--- {cls_threats}\n")

    for prp, hdr in zip(['text', 'sum'], ['full_text_bottom_up_threats', 'summary_bottom_up_threats']):
        text = output[prp]
        cls_threats = output[hdr]

        threats = []
        for k1, v1 in category_dict.items():
            if isinstance(v1, dict):
                for k1, v2 in v1.items():
                    threats.extend(v2)
            else:
                threats.extend(v1)

        threat_list = classify_text(llm_chain, text, threats)
        threat_set = set(threat_list)
    
        for k1, v1 in category_dict.items():
            if isinstance(v1, dict):
                for k2, v2 in v1.items():
                    i = list(set(v2).intersection(threat_set))
                    if i:
                        cls_threats.extend([k1, k2] + i)
            else:
                i =  list(set(v1).intersection(threat_set))
                if i:
                    cls_threats.extend([k1] + i)

        print(f"\t[{model_name} {prp} {hdr}] --- {output['url']}\n\t--- {cls_threats}\n")

    return output


def compute_scores(output):
    truths = output['ground_truth']
    output_scores = defaultdict(list)
    for hdr in ['full_text_top_down_threats', 'summary_top_down_threats', 'full_text_bottom_up_threats', 'summary_bottom_up_threats']:
        prediction = output[hdr]
        true_positives = len(set(prediction).intersection(set(truths)))
        false_positives = len(set(prediction).difference(set(truths)))
        true_negatives = 0
        false_negatives = len(set(truths).difference(set(prediction)))
        # print(f"[{hdr}] {true_positives} {false_positives} {true_negatives} {false_negatives}")
        if true_positives > 0:
            precision = true_positives * 1.0 / (true_positives + false_positives)
            recall = true_positives * 1.0 / (true_positives + false_negatives)
        else:
            precision = 0
            if true_positives + false_positives > 0:
                precision = true_positives * 1.0 / (true_positives + false_positives)
            recall = 0
            if true_positives + false_negatives == 0:
                recall = true_positives * 1.0 / (true_positives + false_negatives)

        score_hdr = hdr.replace('_threats', '_f1_score')
        output[score_hdr] = 0
        if precision + recall > 0:
            output[score_hdr] = 2 * (precision * recall) / (precision + recall) 
        output_scores[output[score_hdr]].append(score_hdr)
            
        score_hdr = hdr.replace('_threats', '_fh_score')
        output[score_hdr] = 0
        if precision + recall > 0:
            output[score_hdr] = (1.25) * (precision * recall) / (0.25 * precision + recall)
        output_scores[output[score_hdr]].append(score_hdr)

    for score in sorted(output_scores.keys(), reverse=True):
        output['best_scores'].append([score, output_scores[score]])
    return output


if __name__ == '__main__':
    config_file_name = sys.argv[1]
    input_data_file = sys.argv[2]
    input_sheet_name = sys.argv[3]
    llm_model_name = sys.argv[4]

    config = load_config(config_file_name)
    categories = load_categories(config)
    inputs = load_data_file(input_data_file, input_sheet_name, categories)
    
    output_parser = NumberedListOutputParser()
    format_instructions = output_parser.get_format_instructions()

    llm_model = Ollama(model=llm_model_name, temperature=0.0)

    summarizer_prompt = create_summarizer_prompt(format_instructions)
    sum_llm_chain = summarizer_prompt | llm_model | output_parser

    classifier_prompt = create_classifier_prompt(format_instructions)
    cls_llm_chain = classifier_prompt | llm_model | output_parser

    outputs = []
    best_methods = dict()
    for input in inputs:
        if not input['url']:
            continue

        output = {
            'url': input['url'],
            'text': input['text'],
            'sum': '',
            'ground_truth': input['ground_truth'],
            'full_text_top_down_threats': [],
            'full_text_bottom_up_threats': [],
            'summary_top_down_threats': [],
            'summary_bottom_up_threats': [],
            'full_text_top_down_f1_score': [],
            'full_text_top_down_fh_score': [],
            'full_text_bottom_up_f1_score': [],
            'full_text_bottom_up_fh_score': [],
            'summary_top_down_f1_score': [],
            'summary_top_down_fh_score': [],
            'summary_bottom_up_f1_score': [],
            'summary_bottom_up_fh_score': [],
            'best_scores': [],
            'best_methods': dict(),
        }

        output = summarize_text(llm_model_name, sum_llm_chain, output)
        output = classify_threats(llm_model_name, cls_llm_chain, categories, output)
        output = compute_scores(output)
        outputs.append(output)

        for i, score_headers in enumerate(output['best_scores']):
            score, headers = score_headers
            for header in headers:
                output['best_methods'][header] = [i, score]
                if header not in best_methods:
                    best_methods[header] = [0 for i in range(0, 8)]
                best_methods[header][i] += 1

        print(f"{output['url']} --- {output['best_scores']} --- {output['best_methods']}\n")
        print(f"{best_methods}\n")

    output_excel_file = input_data_file.replace(".xlsx", f"-{llm_model_name}-out.xlsx")
    write_to_excel_file(output_excel_file, 'Output', outputs)
